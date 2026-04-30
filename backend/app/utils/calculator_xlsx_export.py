"""Render a CalculatorConfig.payload as a human-readable xlsx workbook.

The format is intentionally not a round-trip of :class:`CalculatorExtractor` —
it's an operator-friendly snapshot grouped by scenario. Each section uses
Russian column headers and bold row labels so the operator can edit the
file in Excel without referring to the schema.

The xlsx isn't intended to be re-imported as-is; for editing-and-importing
the operator should use the source template the developer provided. Once
we ship a strict bidirectional schema we'll add a separate "Шаблон для
импорта" download.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill(start_color="222B36", end_color="222B36", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SECTION_FILL = PatternFill(start_color="E14B3A", end_color="E14B3A", fill_type="solid")
_SECTION_FONT = Font(color="FFFFFF", bold=True, size=12)


_SCENARIO_LABELS = {
    "under_3": "До 3 лет",
    "3_5": "3-5 лет",
    "electric": "Электро / Гибрид",
}


_EXPENSE_LABELS = {
    "bank": "Банк, за перевод",
    "purchase": "Покупка по НЕТТО",
    "inspection": "Осмотр подборщиком",
    "delivery_eu_minsk": "Доставка Европа → Минск",
    "delivery_eu_msk": "Доставка Европа → Москва",
    "delivery_msk": "Доставка Минск → Москва",
    "customs_by": "Таможня РБ",
    "transfer_fee": "Комиссия за перевод денег за таможню",
    "elpts": "ЭЛПТС",
    "insurance": "Страхование, брокер",
    "investor": "Инвестор",
    "broker_elpts": "Брокер и ЭлПТС",
    "customs_fee": "Таможенный сбор",
}


def _section_row(ws, title: str) -> None:
    row = ws.max_row + 2
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = _SECTION_FONT
    cell.fill = _SECTION_FILL
    cell.alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)


def _header_row(ws, headers: List[str]) -> None:
    row = ws.max_row + 1
    for idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=idx, value=value)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _table_rows(ws, rows: List[List[Any]]) -> None:
    for r in rows:
        ws.append(r)


def _autosize_columns(ws, max_col: int = 6) -> None:
    for c in range(1, max_col + 1):
        max_len = 12
        for cell in ws[get_column_letter(c)]:
            try:
                length = len(str(cell.value)) if cell.value is not None else 0
            except Exception:
                length = 0
            if length > max_len:
                max_len = length
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 40)


def _expenses_section(ws, expenses: Dict[str, Any]) -> None:
    if not expenses:
        return
    _section_row(ws, "Расходы")
    _header_row(ws, ["Статья", "Сумма (₽)", "Ключ"])
    for key, value in expenses.items():
        label = _EXPENSE_LABELS.get(key, key)
        ws.append([label, value, key])


def _table_section(
    ws,
    title: str,
    headers: List[str],
    rows: List[List[Any]],
) -> None:
    if not rows:
        return
    _section_row(ws, title)
    _header_row(ws, headers)
    _table_rows(ws, rows)


def render_calculator_payload(payload: Dict[str, Any]) -> bytes:
    """Return xlsx bytes that summarise ``payload``.

    The layout has four sheets: Сводка, До 3 лет, 3-5 лет, Электро.
    """

    wb = Workbook()

    # ── Summary sheet
    summary = wb.active
    summary.title = "Сводка"
    meta = payload.get("meta") or {}
    summary.append(["Курс EUR (по умолчанию)", meta.get("eur_rate_default")])
    summary.append(["Курс USD (по умолчанию)", meta.get("usd_rate_default")])
    summary.append([])
    summary.append(["Сценарий", "Документация"])
    for label in _SCENARIO_LABELS.values():
        summary.append([label, f"См. лист «{label}»"])
    _autosize_columns(summary, 4)

    scenarios = payload.get("scenarios") or {}

    # ── Under-3 sheet
    under3 = scenarios.get("under_3") or {}
    sheet = wb.create_sheet(_SCENARIO_LABELS["under_3"])
    _expenses_section(sheet, under3.get("expenses") or {})
    _table_section(
        sheet,
        "Таможенная пошлина по объёму, EUR/см³",
        ["От, см³", "До, см³", "EUR за см³"],
        [[r.get("from"), r.get("to"), r.get("eur_per_cc")] for r in under3.get("duty_by_cc") or []],
    )
    _table_section(
        sheet,
        "Утиль-сбор по объёму",
        ["От, см³", "До, см³", "Сумма (₽)"],
        [[r.get("from"), r.get("to"), r.get("rub")] for r in under3.get("util_by_cc") or []],
    )
    _autosize_columns(sheet, 6)

    # ── 3-5 sheet
    three5 = scenarios.get("3_5") or {}
    sheet = wb.create_sheet(_SCENARIO_LABELS["3_5"])
    _expenses_section(sheet, three5.get("expenses") or {})
    _table_section(
        sheet,
        "Таможенная пошлина по объёму, EUR/см³",
        ["От, см³", "До, см³", "EUR за см³"],
        [[r.get("from"), r.get("to"), r.get("eur_per_cc")] for r in three5.get("duty_by_cc") or []],
    )
    _table_section(
        sheet,
        "Утиль-сбор по объёму",
        ["От, см³", "До, см³", "Сумма (₽)"],
        [[r.get("from"), r.get("to"), r.get("rub")] for r in three5.get("util_by_cc") or []],
    )
    _section_row(sheet, "Дополнительно")
    sheet.append(["Таможенный сбор (₽)", three5.get("customs_fee_rub")])
    sheet.append(["Брокер + ЭлПТС (₽)", three5.get("broker_elpts_rub")])
    _autosize_columns(sheet, 6)

    # ── Electric sheet
    electric = scenarios.get("electric") or {}
    sheet = wb.create_sheet(_SCENARIO_LABELS["electric"])
    _expenses_section(sheet, electric.get("expenses") or {})
    _section_row(sheet, "Параметры")
    sheet.append(["Пошлина, %", electric.get("duty_percent")])
    sheet.append(["НДС, %", electric.get("vat_percent")])
    sheet.append(["Утиль-сбор (₽)", electric.get("util_rub")])
    sheet.append(["Таможенный сбор (₽)", electric.get("customs_fee_rub")])
    sheet.append(["Брокер + ЭлПТС (₽)", electric.get("broker_elpts_rub")])
    _table_section(
        sheet,
        "Акциз по л.с.",
        ["От, л.с.", "До, л.с.", "₽ за л.с."],
        [[r.get("from_hp"), r.get("to_hp"), r.get("rub_per_hp")] for r in electric.get("excise_by_hp") or []],
    )
    _table_section(
        sheet,
        "Сбор за мощность по возрасту",
        ["От, л.с.", "До, л.с.", "Возраст", "Сумма (₽)"],
        [
            [r.get("from_hp"), r.get("to_hp"), r.get("age_bucket"), r.get("rub")]
            for r in electric.get("power_fee") or []
        ],
    )
    _autosize_columns(sheet, 6)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
