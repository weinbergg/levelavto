from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Dict, Any, Iterable, List, Tuple
import datetime as dt
import yaml


CUSTOMS_PATH = Path("/app/backend/app/config/customs.yml")


@dataclass
class ApplyStats:
    updated: int = 0
    skipped: int = 0
    errors: int = 0


def _resolve_customs_path() -> Path:
    if CUSTOMS_PATH.exists():
        return CUSTOMS_PATH
    return Path(__file__).resolve().parent.parent / "config" / "customs.yml"


def load_customs_dict() -> Dict[str, Any]:
    path = _resolve_customs_path()
    data = yaml.safe_load(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError("customs.yml invalid or empty")
    return data


def save_customs_dict(data: Dict[str, Any]) -> None:
    path = _resolve_customs_path()
    path.write_text(
        yaml.safe_dump(data, allow_unicode=True, sort_keys=False),
        encoding="utf-8",
    )


def bump_customs_version(data: Dict[str, Any]) -> None:
    today = dt.datetime.now().strftime("%Y_%m_%d")
    data["version"] = str(today)


def _iter_tables(data: Dict[str, Any], age_bucket: str) -> Dict[str, Any]:
    key = {
        "under_3": "util_tables_under3",
        "3_5": "util_tables_3_5",
        "electric": "util_tables_electric",
    }.get(age_bucket, "util_tables")
    tables = data.get(key) or data.get("util_tables") or {}
    if not isinstance(tables, dict):
        return {}
    return tables


_AGE_LABELS_RU = {
    "under_3": "ДО 3 ЛЕТ (новые автомобили)",
    "3_5": "3–5 ЛЕТ (бывшие в употреблении)",
    "electric": "ЭЛЕКТРО / ГИБРИД (любой возраст)",
}

_TABLE_LABELS_RU = {
    "personal": "Физ. лица (для личного пользования)",
    "commercial": "Юр. лица / коммерческое использование",
    "import": "Юр. лица / импорт под перепродажу",
}

_POWER_LABELS_RU = {
    "kw": "кВт",
    "hp": "л.с.",
}


def _table_label_ru(name: str) -> str:
    return _TABLE_LABELS_RU.get(name, name)


def build_util_template(data: Dict[str, Any]) -> str:
    """Build a TXT template that the operator's bookkeeper can edit in Excel.

    The schema (5-tuple key + price) is kept the same — :func:`apply_util_template`
    parses it back — but each section gets a human-readable Russian heading
    so the operator does not need to remember what ``under_3,personal,kw,…``
    means.
    """

    lines: List[str] = [
        "# ════════════════════════════════════════════════════════════════════",
        "# Шаблон ставок утилизационного сбора",
        "#",
        "# Что заполнять:",
        "#   1. Меняйте только ПОСЛЕДНЕЕ число в строке (сумма в рублях).",
        "#   2. Не трогайте первые пять полей — это идентификатор ставки.",
        "#   3. Любая строка, начинающаяся с #, — комментарий, она игнорируется.",
        "#   4. Можно открыть в Excel: разделители — запятая, точка с запятой",
        "#      или табуляция. Файл не зашифрован, можно править блокнотом.",
        "#",
        "# Формат строки (всего 6 значений):",
        "#   age_bucket , cc_table , power_type , от , до , сумма_руб",
        "#",
        "# Возможные значения:",
        "#   age_bucket  : under_3   = до 3 лет (новые)",
        "#                 3_5       = 3–5 лет (б/у)",
        "#                 electric  = электро/гибрид",
        "#   cc_table    : название категории (personal / commercial / …) ",
        "#                 — задаётся в customs.yml, новые названия не появятся",
        "#                 при загрузке этого файла, добавляйте их через ИТ.",
        "#   power_type  : kw  = киловатты",
        "#                 hp  = лошадиные силы",
        "#",
        "# После загрузки изменения попадают в каталог через 1–2 минуты",
        "# (сбрасывается кэш и пересчитывается стоимость).",
        "# ════════════════════════════════════════════════════════════════════",
        "",
    ]

    for age_bucket in ("under_3", "3_5", "electric"):
        tables = _iter_tables(data, age_bucket)
        if not tables:
            continue
        lines.append("")
        lines.append(f"# ─── {_AGE_LABELS_RU.get(age_bucket, age_bucket)} ───")
        for table_name, table in tables.items():
            lines.append(f"#   • Категория: {_table_label_ru(table_name)}")
            for power_type in ("kw", "hp"):
                rows = (table or {}).get(power_type) or []
                if not rows:
                    continue
                lines.append(f"#     ↳ Мощность в {_POWER_LABELS_RU.get(power_type, power_type)}:")
                for row in rows:
                    lines.append(
                        f"{age_bucket},{table_name},{power_type},"
                        f"{row.get('from', 0)},{row.get('to')},{row.get('price_rub')}"
                    )
    lines.append("")
    return "\n".join(lines)


def _strip_inline_comment(line: str) -> str:
    """Drop any ``# …`` tail so operators can annotate rows in the file."""

    hash_idx = line.find("#")
    if hash_idx == -1:
        return line.strip()
    return line[:hash_idx].strip()


def _split_line(line: str) -> List[str]:
    # allow comma, semicolon, or tab
    for sep in (",", ";", "\t"):
        parts = [p.strip() for p in line.split(sep)]
        if len(parts) >= 6:
            return parts
    return [p.strip() for p in line.split(",")]


def build_util_xlsx(data: Dict[str, Any]) -> bytes:
    """Build an Excel workbook with util fees split by age bucket.

    Layout:
      • One sheet per age bucket: ``ДО 3 ЛЕТ`` / ``3–5 ЛЕТ`` /
        ``ЭЛЕКТРО · ГИБРИД``.
      • Each sheet has section blocks per cc_table (`personal`,
        `commercial`, `import` …), inside it — sub-tables per power_type
        (`kw` / `hp`).
      • Column layout: Категория | Тип мощности | От | До | Сумма (₽).
        Operator edits ONLY the «Сумма (₽)» column (right-most).
      • The first row of every cc_table block is a colored header so the
        bookkeeper can navigate the file without remembering the schema.

    The result is symmetric with :func:`apply_util_xlsx`: download →
    edit in Excel → upload → applied. No hidden columns, no formulas —
    just the same 5 fields that already exist in customs.yml.
    """

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="222B36", end_color="222B36", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    section_fill = PatternFill(start_color="E14B3A", end_color="E14B3A", fill_type="solid")
    section_font = Font(color="FFFFFF", bold=True, size=12)
    sub_fill = PatternFill(start_color="F1B400", end_color="F1B400", fill_type="solid")
    sub_font = Font(color="222B36", bold=True)

    wb = Workbook()

    # Drop the default empty sheet — we'll create one per age bucket.
    default_sheet = wb.active
    wb.remove(default_sheet)

    sheet_names = {
        "under_3": "ДО 3 ЛЕТ",
        "3_5": "3–5 ЛЕТ",
        "electric": "ЭЛЕКТРО ГИБРИД",
    }

    for age_bucket in ("under_3", "3_5", "electric"):
        tables = _iter_tables(data, age_bucket)
        ws = wb.create_sheet(sheet_names.get(age_bucket, age_bucket))

        # ── Top intro block ──────────────────────────────────────
        intro = ws.cell(row=1, column=1, value=_AGE_LABELS_RU.get(age_bucket, age_bucket))
        intro.font = section_font
        intro.fill = section_fill
        intro.alignment = Alignment(vertical="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        ws.row_dimensions[1].height = 24

        ws.cell(row=2, column=1, value=(
            "Меняйте только колонку «Сумма (₽)». "
            "Не редактируйте «Категория», «Тип мощности», «От», «До» — это идентификатор ставки."
        )).font = Font(italic=True, color="555555")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

        if not tables:
            ws.cell(row=4, column=1, value="(нет ставок в конфиге)").font = Font(italic=True)
            for col in range(1, 6):
                ws.column_dimensions[get_column_letter(col)].width = 22
            continue

        # ── Column headers ───────────────────────────────────────
        headers = ["Категория", "Тип мощности", "От", "До", "Сумма (₽)"]
        for idx, value in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=idx, value=value)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        cur_row = 5
        for table_name, table in tables.items():
            label = _table_label_ru(table_name)
            # Sub-section header — visually separates personal / commercial.
            scell = ws.cell(row=cur_row, column=1, value=f"▶ {label}")
            scell.font = sub_font
            scell.fill = sub_fill
            scell.alignment = Alignment(vertical="center")
            ws.merge_cells(
                start_row=cur_row, start_column=1, end_row=cur_row, end_column=5
            )
            cur_row += 1

            for power_type in ("kw", "hp"):
                rows = (table or {}).get(power_type) or []
                if not rows:
                    continue
                for row in rows:
                    ws.cell(row=cur_row, column=1, value=table_name)
                    ws.cell(row=cur_row, column=2, value=power_type)
                    ws.cell(row=cur_row, column=3, value=row.get("from", 0))
                    ws.cell(row=cur_row, column=4, value=row.get("to"))
                    price_cell = ws.cell(row=cur_row, column=5, value=row.get("price_rub"))
                    price_cell.font = Font(bold=True)
                    cur_row += 1

        # ── Column widths ────────────────────────────────────────
        widths = [22, 14, 12, 12, 18]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        # Freeze the header row so scrolling long tables stays readable.
        ws.freeze_panes = "A5"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def apply_util_xlsx(data: Dict[str, Any], file_bytes: bytes) -> ApplyStats:
    """Parse the operator-edited xlsx (built by :func:`build_util_xlsx`)
    and write changed prices back into ``data`` in-place.

    Recognises the layout produced above: one sheet per age bucket, the
    sheet name decides which ``util_tables_*`` key is updated. Inside
    the sheet we look at columns:
      [Категория, Тип мощности, От, До, Сумма (₽)]
    Section/sub-section header rows (merged or non-numeric) are silently
    skipped.

    The operator can keep or rename sheets — what matters is the first
    line being one of the recognised intro labels (``ДО 3 ЛЕТ``, etc.).
    Values that fail to parse don't crash the import; they bump the
    ``errors`` counter so the operator sees the diagnostic in the flash.
    """

    from openpyxl import load_workbook

    stats = ApplyStats()

    sheet_to_age = {
        "ДО 3 ЛЕТ": "under_3",
        "3-5 ЛЕТ": "3_5",
        "3–5 ЛЕТ": "3_5",
        "ЭЛЕКТРО / ГИБРИД": "electric",
        "ЭЛЕКТРО ГИБРИД": "electric",
        "ЭЛЕКТРО": "electric",
    }

    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception:
        stats.errors += 1
        return stats

    for ws in wb.worksheets:
        # Match by sheet name first, fall back to the merged intro cell.
        age_bucket = sheet_to_age.get(ws.title.strip().upper())
        if not age_bucket:
            intro = (ws.cell(row=1, column=1).value or "")
            for label, bucket in sheet_to_age.items():
                if label in str(intro).upper():
                    age_bucket = bucket
                    break
        if not age_bucket:
            stats.skipped += 1
            continue

        tables_key = {
            "under_3": "util_tables_under3",
            "3_5": "util_tables_3_5",
            "electric": "util_tables_electric",
        }.get(age_bucket, "util_tables")
        tables = data.get(tables_key)
        if tables is None:
            tables = {}
            data[tables_key] = tables

        # Iterate from the first data row onwards. Header is row 4 in
        # the canonical layout, but we skip any row whose first cell is
        # not a recognisable category — that gracefully ignores intro
        # lines / sub-section headers / blank rows.
        for raw_row in ws.iter_rows(min_row=2, values_only=True):
            if not raw_row or len(raw_row) < 5:
                continue
            cat, power_type, from_v, to_v, price_v = (
                raw_row[0], raw_row[1], raw_row[2], raw_row[3], raw_row[4]
            )
            cat_s = (str(cat).strip() if cat is not None else "")
            pt_s = (str(power_type).strip().lower() if power_type is not None else "")
            if pt_s not in ("kw", "hp"):
                continue
            if not cat_s or cat_s.startswith("▶") or cat_s.lower() in ("категория",):
                continue
            try:
                target_from = float(from_v)
                target_to = float(to_v)
                target_price = float(price_v)
            except (TypeError, ValueError):
                stats.skipped += 1
                continue

            table = tables.get(cat_s)
            if table is None:
                tables[cat_s] = {"kw": [], "hp": []}
                table = tables[cat_s]
            rows = table.setdefault(pt_s, [])

            updated = False
            for row in rows:
                if (
                    float(row.get("from", 0)) == target_from
                    and float(row.get("to")) == target_to
                ):
                    row["price_rub"] = target_price
                    updated = True
                    break
            if not updated:
                rows.append({
                    "from": target_from,
                    "to": target_to,
                    "price_rub": target_price,
                })
            stats.updated += 1

    return stats


def apply_util_template(data: Dict[str, Any], content: str) -> ApplyStats:
    stats = ApplyStats()
    cleaned_lines: List[str] = []
    for raw in content.splitlines():
        stripped = _strip_inline_comment(raw)
        if stripped:
            cleaned_lines.append(stripped)
    for ln in cleaned_lines:
        try:
            parts = _split_line(ln)
            if len(parts) < 6:
                stats.skipped += 1
                continue
            age_bucket, table_name, power_type, from_s, to_s, price_s = parts[:6]
            age_bucket = age_bucket.strip()
            power_type = power_type.strip()
            if power_type not in ("kw", "hp"):
                stats.skipped += 1
                continue
            tables_key = {
                "under_3": "util_tables_under3",
                "3_5": "util_tables_3_5",
                "electric": "util_tables_electric",
            }.get(age_bucket, "util_tables")
            tables = data.get(tables_key)
            if tables is None:
                tables = {}
                data[tables_key] = tables
            table = tables.get(table_name)
            if table is None:
                tables[table_name] = {"kw": [], "hp": []}
                table = tables[table_name]
            rows = table.get(power_type)
            if rows is None:
                table[power_type] = []
                rows = table[power_type]
            target_from = float(from_s)
            target_to = float(to_s)
            target_price = float(price_s)
            updated = False
            for row in rows:
                if float(row.get("from", 0)) == target_from and float(row.get("to")) == target_to:
                    row["price_rub"] = target_price
                    updated = True
                    break
            if not updated:
                rows.append({"from": target_from, "to": target_to, "price_rub": target_price})
            stats.updated += 1
        except Exception:
            stats.errors += 1
    return stats

