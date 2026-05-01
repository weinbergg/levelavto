from urllib.parse import quote
from datetime import datetime, timedelta
import json
import logging
import secrets

from fastapi import APIRouter, Depends, Form, Request, UploadFile, File
from fastapi.responses import RedirectResponse, Response
from sqlalchemy.orm import Session
from sqlalchemy import select, func
from pathlib import Path
import tempfile

from ..auth import get_current_user, require_admin
from ..db import get_db
from ..services.admin_service import AdminService
from ..services.cars_service import CarsService
from ..services.calculator_config_service import CalculatorConfigService
from ..services.calculator_extractor import CalculatorExtractor
from ..services.content_service import ContentService
from ..services.customs_config import reset_customs_config_cache
from ..utils.recommended_config import load_config, save_config, DEFAULT_CONFIG
from ..utils.home_content import build_home_content, default_home_content, serialize_home_content
from ..utils.customs_template import (
    load_customs_dict,
    save_customs_dict,
    bump_customs_version,
    build_util_template,
    apply_util_template,
)
from ..utils.brand_groups import (
    BRAND_FILTER_PRIORITY,
    TOP_BRANDS_CONTENT_KEY,
    TOP_MODELS_CONTENT_KEY,
    _coerce_models_override,
    _coerce_priority_list,
    _normalize_to_priority,
    effective_priority,
    load_models_priority,
)
from ..utils.redis_cache import bump_dataset_version, redis_delete_by_pattern
from ..models import Car, CalculatorConfig, Favorite, Notification, PageVisit, User
from ..services.notification_service import NotificationService
import time


router = APIRouter()
logger = logging.getLogger(__name__)


def _admin_redirect(message: str = "", *, error: str = "", path: str = "/admin") -> RedirectResponse:
    """Build a 302 redirect back to the admin with a toast hint."""
    parts = []
    if error:
        parts.append("flash_error=" + quote(error))
    if message:
        parts.append("flash=" + quote(message))
    suffix = ("?" + "&".join(parts)) if parts else ""
    return RedirectResponse(url=path + suffix, status_code=302)

CONTACT_KEYS = {
    "contact_phone": "Телефон",
    "contact_email": "Email",
    "contact_address": "Адрес",
    "contact_tg": "Telegram",
    "contact_max": "MAX",
    "contact_vk": "VK",
    "contact_avito": "Avito",
    "contact_autoru": "Auto.ru",
    "contact_wa": "WhatsApp",
    "contact_ig": "Instagram",
    "contact_map_link": "Ссылка на карту",
    "lead_email": "Email для заявок",
}
LEGACY_HOME_KEYS = ("hero_title", "hero_subtitle", "hero_note")


@router.get("/admin")
def admin_dashboard(
    request: Request,
    user: User | None = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    # On the HTML entry-point we replace the generic 401/403 JSON errors
    # with friendly redirects so the operator gets a login form instead of
    # a raw {"detail": …} response when the session has expired or when a
    # non-admin user lands on /admin by mistake.
    if user is None:
        return RedirectResponse(url="/login?next=" + quote("/admin"), status_code=302)
    if not user.is_admin:
        return RedirectResponse(
            url="/?" + "flash_error=" + quote("Доступ к админке только у администраторов"),
            status_code=302,
        )
    templates = request.app.state.templates
    admin_svc = AdminService(db)
    cars_svc = CarsService(db)
    featured_recommended = admin_svc.list_featured("recommended")
    content = admin_svc.list_site_content(
        list(CONTACT_KEYS.keys()) + ["home_content", *LEGACY_HOME_KEYS]
    )
    home_content = build_home_content(content)
    recommended_cfg = load_config()
    calc_cfg = CalculatorConfigService(db).latest()
    total_cars = cars_svc.total_cars()
    overview_stats = admin_svc.overview_stats()
    customs_data = load_customs_dict()
    cc_tables = set()
    for key in ("util_tables_under3", "util_tables_3_5", "util_tables_electric", "util_tables"):
        tables = customs_data.get(key) or {}
        if isinstance(tables, dict):
            cc_tables.update(tables.keys())
    customs_cc_tables = sorted(cc_tables)
    return templates.TemplateResponse(
        "admin/dashboard.html",
        {
            "request": request,
            "user": user,
            "content": content,
            "home": home_content,
            "recommended_cfg": recommended_cfg,
            "featured_recommended": featured_recommended,
            "calc_cfg": calc_cfg,
            "total_cars": total_cars,
            "overview_stats": overview_stats,
            "customs_cc_tables": customs_cc_tables,
            "page_title": "Сводка",
            "page_subtitle": "Общая статистика и быстрый доступ к разделам админки.",
            "breadcrumbs": [("Админка", None)],
        },
    )


def _render_coming_soon(
    request: Request,
    user: User,
    *,
    page_title: str,
    page_subtitle: str,
    stub_title: str,
    stub_description: str,
    stub_features: list[str] | None = None,
    stub_eta: str = "",
    breadcrumbs: list[tuple[str, str | None]] | None = None,
):
    templates = request.app.state.templates
    return templates.TemplateResponse(
        "admin/_coming_soon.html",
        {
            "request": request,
            "user": user,
            "page_title": page_title,
            "page_subtitle": page_subtitle,
            "stub_title": stub_title,
            "stub_description": stub_description,
            "stub_features": stub_features or [],
            "stub_eta": stub_eta,
            "breadcrumbs": breadcrumbs
            or [("Админка", "/admin"), (page_title, None)],
        },
    )


def _all_known_brands(db: Session) -> list[str]:
    """Return the full list of brand names — DB facets ∪ default priority.

    The default priority brands are always present even when no inventory
    matches them yet (e.g. Lamborghini after a slow week), so the operator
    can keep them in the "top" group. DB facets give the long tail.
    """

    rows = db.execute(
        select(func.coalesce(Car.brand, "")).where(Car.brand.isnot(None)).distinct()
    ).all()
    seen: dict[str, str] = {}
    for (raw,) in rows:
        text = str(raw or "").strip()
        if not text:
            continue
        canon = _normalize_to_priority(text) or text
        key = canon.casefold()
        if key not in seen:
            seen[key] = canon
    for canon in BRAND_FILTER_PRIORITY:
        key = canon.casefold()
        if key not in seen:
            seen[key] = canon
    return sorted(seen.values(), key=lambda v: v.casefold())


def _bump_filter_caches() -> None:
    """Drop any cached filter contexts and bump the dataset version.

    Called after editing the operator-controlled top-brands list so the
    very next request rebuilds ``filter_ctx_base`` with the new ordering.
    Also nukes the in-process caches inside the pages router — without
    this, every gunicorn worker keeps serving stale data from its own
    TTLCache for up to 15 minutes (which is exactly what the operator
    saw when changes "didn't apply" even in incognito).
    """

    try:
        redis_delete_by_pattern("filter_ctx_*")
        redis_delete_by_pattern("home_*")
        bump_dataset_version()
    except Exception:
        logger.exception("admin: failed to bump filter caches after top-brands edit")
    _drop_pages_in_process_caches()


def _bump_home_caches() -> None:
    """Drop home-page caches after operator changes featured/recommended."""

    try:
        redis_delete_by_pattern("home_*")
    except Exception:
        logger.exception("admin: failed to drop home_* redis keys")
    _drop_pages_in_process_caches()


def _drop_pages_in_process_caches() -> None:
    """Best-effort wipe of the per-worker TTLCaches inside pages.py.

    Each gunicorn worker keeps its own LRU/TTL caches for the home page
    and for the public filter context. Bumping ``dataset_version`` is
    not enough because some of those keys do not embed the version yet,
    so we explicitly call ``.clear()`` on every cache the pages module
    exposes. This only affects the worker handling this request, but
    Redis-side invalidation above ensures every other worker rebuilds
    on its next request anyway.
    """

    try:
        from . import pages as _pages_router  # local import to avoid cycles
    except Exception:
        logger.exception("admin: cannot import pages router to drop in-process caches")
        return
    for cache_name in (
        "_FILTER_CTX_CACHE",
        "_TOTAL_CARS_CACHE",
        "_HOME_FILTER_CTX_CACHE",
        "_HOME_MEDIA_CACHE",
        "_HOME_RECOMMENDED_CACHE",
        "_HOME_MORE_OFFERS_CACHE",
    ):
        cache = getattr(_pages_router, cache_name, None)
        if cache is not None:
            try:
                cache.clear()
            except Exception:
                logger.warning("admin: failed to clear %s", cache_name, exc_info=True)


@router.get("/admin/top-brands", response_class=None)
def admin_top_brands_page(
    request: Request,
    user: User | None = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user is None:
        return RedirectResponse(url="/login?next=" + quote("/admin/top-brands"), status_code=302)
    if not user.is_admin:
        return RedirectResponse(url="/", status_code=302)

    override_raw = ContentService(db).get(TOP_BRANDS_CONTENT_KEY)
    override_list = _coerce_priority_list(override_raw)
    active = effective_priority(override_raw)
    all_brands = _all_known_brands(db)

    active_keys = {brand.casefold() for brand in active}
    other_brands = [name for name in all_brands if name.casefold() not in active_keys]

    templates = request.app.state.templates
    return templates.TemplateResponse(
        "admin/top_brands.html",
        {
            "request": request,
            "user": user,
            "page_title": "Топ-марки",
            "page_subtitle": (
                "Перетащите бренды между колонками или поменяйте порядок. "
                "Изменения применятся ко всем формам поиска сразу после сохранения."
            ),
            "breadcrumbs": [
                ("Админка", "/admin"),
                ("Топ-марки", None),
            ],
            "top_brands": active,
            "other_brands": other_brands,
            "default_brands": list(BRAND_FILTER_PRIORITY),
            "is_overridden": override_list is not None,
        },
    )


@router.post("/admin/top-brands")
def admin_top_brands_save(
    request: Request,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
    brands: str = Form(""),
    reset: str = Form(""),
):
    content = ContentService(db)
    if reset == "1":
        content.upsert_content(TOP_BRANDS_CONTENT_KEY, "", description="top brands override")
        _bump_filter_caches()
        return _admin_redirect(
            "Список сброшен — используется набор по умолчанию",
            path="/admin/top-brands",
        )

    requested = [item.strip() for item in brands.split(",") if item.strip()]
    cleaned = _coerce_priority_list(requested)
    if not cleaned:
        return _admin_redirect(
            error="Нужно выбрать хотя бы один бренд",
            path="/admin/top-brands",
        )
    content.upsert_content(
        TOP_BRANDS_CONTENT_KEY,
        json.dumps(cleaned, ensure_ascii=False),
        description="top brands override",
    )
    _bump_filter_caches()
    return _admin_redirect(
        f"Список сохранён ({len(cleaned)} марок)",
        path="/admin/top-brands",
    )


@router.get("/admin/top-models", response_class=None)
def admin_top_models_page(
    request: Request,
    user: User | None = Depends(get_current_user),
    db: Session = Depends(get_db),
    brand: str = "",
):
    if user is None:
        return RedirectResponse(url="/login?next=" + quote("/admin/top-models"), status_code=302)
    if not user.is_admin:
        return RedirectResponse(url="/", status_code=302)

    overrides = load_models_priority(db)
    all_brands = _all_known_brands(db)

    brand_pick = (brand or "").strip()
    selected_brand = None
    selected_groups: list[dict] = []
    selected_priority: list[str] = []
    if brand_pick:
        for candidate in all_brands:
            if candidate.casefold() == brand_pick.casefold():
                selected_brand = candidate
                break
        if selected_brand is None:
            selected_brand = brand_pick
        cars_svc = CarsService(db)
        models = cars_svc.models_for_brand_filtered(
            region=None, country=None, kr_type=None, brand=selected_brand,
        )
        groups = cars_svc.build_model_groups(brand=selected_brand, models=models)
        selected_groups = [
            {"label": g.get("label") or "", "count": int(g.get("count") or 0)}
            for g in groups if g.get("label")
        ]
        selected_priority = list(overrides.get(selected_brand.casefold(), []))

    templates = request.app.state.templates
    return templates.TemplateResponse(
        "admin/top_models.html",
        {
            "request": request,
            "user": user,
            "page_title": "Топ-модели",
            "page_subtitle": (
                "Перетащите модели в левую колонку, чтобы они появлялись "
                "первыми в фильтре «Модель» для выбранной марки."
            ),
            "breadcrumbs": [
                ("Админка", "/admin"),
                ("Топ-модели", None),
            ],
            "all_brands": all_brands,
            "selected_brand": selected_brand,
            "selected_groups": selected_groups,
            "selected_priority": selected_priority,
            "overrides_summary": [
                {"brand": brand_name, "count": len(items)}
                for brand_name, items in overrides.items()
                if items
            ],
        },
    )


@router.post("/admin/top-models")
def admin_top_models_save(
    request: Request,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
    brand: str = Form(""),
    models: str = Form(""),
    reset: str = Form(""),
):
    brand_clean = (brand or "").strip()
    if not brand_clean:
        return _admin_redirect(error="Не указана марка", path="/admin/top-models")

    content = ContentService(db)
    current_raw = content.get(TOP_MODELS_CONTENT_KEY)
    overrides = _coerce_models_override(current_raw)
    canonical_keyed: dict[str, list[str]] = {**overrides}

    if reset == "1":
        canonical_keyed.pop(brand_clean.casefold(), None)
    else:
        items = [item.strip() for item in models.split(",") if item.strip()]
        if not items:
            canonical_keyed.pop(brand_clean.casefold(), None)
        else:
            canonical_keyed[brand_clean.casefold()] = items

    save_payload = {brand_name: items for brand_name, items in canonical_keyed.items() if items}

    pretty: dict[str, list[str]] = {}
    for key, items in save_payload.items():
        pretty[brand_clean if key == brand_clean.casefold() else key] = items
    content.upsert_content(
        TOP_MODELS_CONTENT_KEY,
        json.dumps(pretty, ensure_ascii=False),
        description="top models override",
    )
    _bump_filter_caches()
    if reset == "1":
        msg = f"Список моделей сброшен для «{brand_clean}»"
    else:
        msg = f"Сохранено приоритетных моделей для «{brand_clean}»: {len(save_payload.get(brand_clean.casefold(), save_payload.get(brand_clean, [])))}"
    return _admin_redirect(
        msg, path="/admin/top-models?brand=" + quote(brand_clean)
    )


@router.get("/admin/calculator/excel", response_class=None)
def admin_calculator_excel_page(
    request: Request,
    user: User | None = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user is None:
        return RedirectResponse(url="/login?next=" + quote("/admin/calculator/excel"), status_code=302)
    if not user.is_admin:
        return RedirectResponse(url="/", status_code=302)

    svc = CalculatorConfigService(db)
    latest = svc.latest()
    versions = svc.all_versions(limit=10)

    pending = request.session.get("calc_preview")
    pending_summary: dict | None = None
    if pending:
        pending_summary = {
            "filename": pending.get("filename", "файл.xlsx"),
            "total_changes": len(pending.get("changes") or {}),
            "changes": (pending.get("changes") or {}),
            "token": pending.get("token"),
        }

    templates = request.app.state.templates
    return templates.TemplateResponse(
        "admin/calculator_excel.html",
        {
            "request": request,
            "user": user,
            "page_title": "Калькулятор · Excel",
            "page_subtitle": (
                "Скачайте текущий конфиг, отредактируйте в Excel, "
                "загрузите для предпросмотра — изменения применятся только после подтверждения."
            ),
            "breadcrumbs": [
                ("Админка", "/admin"),
                ("Калькулятор · Excel", None),
            ],
            "latest": latest,
            "versions": versions,
            "pending": pending_summary,
        },
    )


@router.get("/admin/calculator/export.xlsx")
def admin_calculator_export_xlsx(
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    from ..utils.calculator_xlsx_export import render_calculator_payload

    latest = CalculatorConfigService(db).latest()
    if not latest:
        return _admin_redirect(error="Конфиг не загружен", path="/admin/calculator/excel")
    data = render_calculator_payload(latest.payload or {})
    filename = f"calculator-config-v{latest.version}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/admin/calculator/preview")
async def admin_calculator_preview(
    request: Request,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
    file: UploadFile = File(...),
):
    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        return _admin_redirect(
            error="Нужен файл .xlsx / .xlsm / .xls",
            path="/admin/calculator/excel",
        )
    try:
        with tempfile.NamedTemporaryFile(delete=False) as tmp:
            tmp.write(await file.read())
            tmp_path = Path(tmp.name)
        payload = CalculatorExtractor(tmp_path).extract()
    except Exception as exc:
        logger.exception("calculator preview parse failed")
        return _admin_redirect(
            error=f"Не удалось прочитать файл: {exc}",
            path="/admin/calculator/excel",
        )

    svc = CalculatorConfigService(db)
    latest = svc.latest()
    changes = svc.diff_payloads(latest.payload if latest else None, payload)
    token = secrets.token_urlsafe(12)
    request.session["calc_preview"] = {
        "token": token,
        "filename": file.filename,
        "payload": payload,
        "changes": changes,
    }
    if not changes:
        return _admin_redirect(
            "Файл прочитан, но он не отличается от текущего конфига",
            path="/admin/calculator/excel",
        )
    return _admin_redirect(
        f"Файл «{file.filename}» проанализирован — {len(changes)} изменений. Проверьте и подтвердите ниже.",
        path="/admin/calculator/excel",
    )


@router.post("/admin/calculator/preview/apply")
def admin_calculator_preview_apply(
    request: Request,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
    token: str = Form(""),
):
    pending = request.session.get("calc_preview")
    if not pending or pending.get("token") != token:
        return _admin_redirect(
            error="Сессия предпросмотра истекла, загрузите файл заново",
            path="/admin/calculator/excel",
        )
    payload = pending.get("payload")
    filename = pending.get("filename") or "preview.xlsx"
    if not payload:
        return _admin_redirect(
            error="Пустой payload — повторите загрузку",
            path="/admin/calculator/excel",
        )
    try:
        CalculatorConfigService(db).create(
            payload=payload,
            source="upload_xlsx_preview",
            comment=f"preview→apply {filename}",
        )
    except Exception as exc:
        logger.exception("calculator preview apply failed")
        return _admin_redirect(
            error=f"Не удалось сохранить конфиг: {exc}",
            path="/admin/calculator/excel",
        )
    request.session.pop("calc_preview", None)
    return _admin_redirect(
        f"Конфиг калькулятора применён (источник: {filename})",
        path="/admin/calculator/excel",
    )


@router.post("/admin/calculator/preview/discard")
def admin_calculator_preview_discard(
    request: Request,
    user: User = Depends(require_admin),
):
    request.session.pop("calc_preview", None)
    return _admin_redirect(
        "Предпросмотр сброшен — изменения не применены",
        path="/admin/calculator/excel",
    )


@router.post("/admin/calculator/rollback")
def admin_calculator_rollback(
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
    version: int = Form(...),
):
    """Roll back to an older calculator config in one click.

    We do not destroy history — instead we copy the chosen version's
    payload into a new row, so the rollback itself becomes auditable.
    """

    svc = CalculatorConfigService(db)
    target = db.execute(
        select(CalculatorConfig).where(CalculatorConfig.version == version)
    ).scalar_one_or_none()
    if target is None:
        return _admin_redirect(
            error=f"Версия v{version} не найдена",
            path="/admin/calculator/excel",
        )
    latest = svc.latest()
    if latest and latest.version == version:
        return _admin_redirect(
            "Эта версия уже актуальна, откат не требуется",
            path="/admin/calculator/excel",
        )
    new_cfg = svc.create(
        payload=target.payload or {},
        source="rollback",
        comment=f"rollback to v{version} (by {user.email})",
    )
    return _admin_redirect(
        f"Откат к v{version} выполнен — теперь актуальна v{new_cfg.version}",
        path="/admin/calculator/excel",
    )


@router.get("/admin/calculator/template.xlsx")
def admin_calculator_template(user: User = Depends(require_admin)):
    """Download an empty (zeroed) calculator template for the customer.

    The structure mirrors the live config, but every numeric value is
    set to 0 / empty so the customer can fill it in from scratch
    without seeing existing rates. Built off the in-code DEFAULT_CONFIG
    skeleton — never the production payload.
    """

    from ..utils.calculator_xlsx_export import render_calculator_payload

    expenses_blank = {
        "bank": 0,
        "purchase": 0,
        "inspection": 0,
        "delivery_eu_minsk": 0,
        "delivery_eu_msk": 0,
        "delivery_msk": 0,
        "customs_by": 0,
        "transfer_fee": 0,
        "elpts": 0,
        "insurance": 0,
        "investor": 0,
        "broker_elpts": 0,
        "customs_fee": 0,
    }
    duty_rows = [
        {"from": 0, "to": 1000, "eur_per_cc": 0},
        {"from": 1000, "to": 1500, "eur_per_cc": 0},
        {"from": 1500, "to": 1800, "eur_per_cc": 0},
        {"from": 1800, "to": 2300, "eur_per_cc": 0},
        {"from": 2300, "to": 3000, "eur_per_cc": 0},
        {"from": 3000, "to": 99999, "eur_per_cc": 0},
    ]
    util_rows = [
        {"from": 0, "to": 1000, "rub": 0},
        {"from": 1000, "to": 2000, "rub": 0},
        {"from": 2000, "to": 3000, "rub": 0},
        {"from": 3000, "to": 99999, "rub": 0},
    ]
    skeleton = {
        "meta": {
            "version": "template",
            "source": "blank",
            "eur_rate_default": 0,
            "usd_rate_default": 0,
        },
        "scenarios": {
            "under_3": {
                "expenses": dict(expenses_blank),
                "duty_by_cc": [dict(row) for row in duty_rows],
                "util_by_cc": [dict(row) for row in util_rows],
            },
            "3_5": {
                "expenses": dict(expenses_blank),
                "duty_by_cc": [dict(row) for row in duty_rows],
                "util_by_cc": [dict(row) for row in util_rows],
                "customs_fee_rub": 0,
                "broker_elpts_rub": 0,
            },
            "electric": {
                "expenses": dict(expenses_blank),
                "duty_percent": 0,
                "vat_percent": 0,
                "util_rub": 0,
                "customs_fee_rub": 0,
                "broker_elpts_rub": 0,
                "excise_by_hp": [
                    {"from_hp": 0, "to_hp": 90, "rub_per_hp": 0},
                    {"from_hp": 90, "to_hp": 150, "rub_per_hp": 0},
                    {"from_hp": 150, "to_hp": 200, "rub_per_hp": 0},
                    {"from_hp": 200, "to_hp": 9999, "rub_per_hp": 0},
                ],
                "power_fee": [],
            },
        },
    }
    data = render_calculator_payload(skeleton)
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="calculator-template-blank.xlsx"'},
    )


@router.post("/admin/calculator/recalc")
def admin_calculator_recalc(
    user: User = Depends(require_admin),
):
    """Kick off the full price-cache rebuild as a background subprocess.

    The actual heavy work happens in ``backend.app.scripts.recalc_calc_cache``
    which already supports retries and chunked commits. We launch it
    detached so the admin page returns instantly; progress can be
    inspected through the container logs (or via the marker file below).
    """

    import subprocess
    import sys

    marker = Path("/tmp/la_recalc_in_progress")
    if marker.exists():
        try:
            age = time.time() - marker.stat().st_mtime
        except FileNotFoundError:
            age = 0
        if age < 7200:  # 2-hour safety window
            return _admin_redirect(
                error="Пересчёт уже запущен. Дождитесь завершения или удалите /tmp/la_recalc_in_progress в контейнере.",
                path="/admin/calculator/excel",
            )

    wrapper = (
        "import sys, pathlib;"
        " sys.argv = ['recalc_calc_cache'];"
        " from backend.app.scripts.recalc_calc_cache import main;"
        " marker = pathlib.Path('/tmp/la_recalc_in_progress');"
        " try:\n    main()\nfinally:\n    marker.unlink(missing_ok=True)"
    )
    try:
        marker.write_text(datetime.utcnow().isoformat())
        subprocess.Popen(  # noqa: S603 — admin-only, no shell interpolation
            [sys.executable, "-c", wrapper],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            close_fds=True,
        )
    except Exception as exc:  # pragma: no cover — surfaces in flash
        logger.exception("recalc launch failed")
        marker.unlink(missing_ok=True)
        return _admin_redirect(
            error=f"Не удалось запустить пересчёт: {exc}",
            path="/admin/calculator/excel",
        )

    return _admin_redirect(
        "Пересчёт запущен в фоне. Это займёт несколько минут — следите за «Машин в базе» на сводке.",
        path="/admin/calculator/excel",
    )


@router.get("/admin/users", response_class=None)
def admin_users_page(
    request: Request,
    user: User | None = Depends(get_current_user),
    db: Session = Depends(get_db),
    q: str = "",
    verified: str = "",
    period: str = "",
    page: int = 1,
):
    if user is None:
        return RedirectResponse(url="/login?next=" + quote("/admin/users"), status_code=302)
    if not user.is_admin:
        return RedirectResponse(url="/", status_code=302)

    page = max(int(page or 1), 1)
    page_size = 25
    offset = (page - 1) * page_size

    fav_count_subq = (
        select(Favorite.user_id, func.count(Favorite.id).label("fav_count"))
        .group_by(Favorite.user_id)
        .subquery()
    )
    base_stmt = select(User, func.coalesce(fav_count_subq.c.fav_count, 0).label("fav_count"))
    base_stmt = base_stmt.outerjoin(
        fav_count_subq, fav_count_subq.c.user_id == User.id
    )
    count_stmt = select(func.count(User.id))

    if q:
        like = f"%{q.strip()}%"
        clause = (
            User.email.ilike(like)
            | User.full_name.ilike(like)
            | User.phone.ilike(like)
        )
        base_stmt = base_stmt.where(clause)
        count_stmt = count_stmt.where(clause)
    if verified == "yes":
        clause = (User.email_verified_at.is_not(None)) | (User.phone_verified_at.is_not(None))
        base_stmt = base_stmt.where(clause)
        count_stmt = count_stmt.where(clause)
    elif verified == "no":
        clause = (User.email_verified_at.is_(None)) & (User.phone_verified_at.is_(None))
        base_stmt = base_stmt.where(clause)
        count_stmt = count_stmt.where(clause)
    if period:
        from datetime import datetime as _dt, timedelta as _td

        days = {"7": 7, "30": 30, "90": 90}.get(period)
        if days:
            since = _dt.utcnow() - _td(days=days)
            base_stmt = base_stmt.where(User.created_at >= since)
            count_stmt = count_stmt.where(User.created_at >= since)

    total = db.execute(count_stmt).scalar_one() or 0
    base_stmt = base_stmt.order_by(User.created_at.desc()).offset(offset).limit(page_size)
    rows = db.execute(base_stmt).all()

    pages = max(1, (total + page_size - 1) // page_size)

    templates = request.app.state.templates
    return templates.TemplateResponse(
        "admin/users.html",
        {
            "request": request,
            "user": user,
            "page_title": "Пользователи",
            "page_subtitle": "Зарегистрированные пользователи, их контакты и активность.",
            "breadcrumbs": [
                ("Админка", "/admin"),
                ("Пользователи", None),
            ],
            "users": [
                {"row": r._mapping["User"], "fav_count": int(r._mapping["fav_count"] or 0)}
                for r in rows
            ],
            "total": total,
            "page": page,
            "pages": pages,
            "page_size": page_size,
            "q": q,
            "verified": verified,
            "period": period,
        },
    )


@router.get("/admin/users/{user_id}", response_class=None)
def admin_user_detail(
    request: Request,
    user_id: int,
    user: User | None = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user is None:
        return RedirectResponse(
            url="/login?next=" + quote(f"/admin/users/{user_id}"),
            status_code=302,
        )
    if not user.is_admin:
        return RedirectResponse(url="/", status_code=302)

    target = db.get(User, user_id)
    if target is None:
        return _admin_redirect(error="Пользователь не найден", path="/admin/users")

    favs = (
        db.execute(
            select(Favorite, Car)
            .join(Car, Car.id == Favorite.car_id)
            .where(Favorite.user_id == user_id)
            .order_by(Favorite.created_at.desc())
        )
        .all()
    )

    templates = request.app.state.templates
    return templates.TemplateResponse(
        "admin/user_detail.html",
        {
            "request": request,
            "user": user,
            "page_title": target.full_name or target.email,
            "page_subtitle": f"Карточка пользователя #{target.id}",
            "breadcrumbs": [
                ("Админка", "/admin"),
                ("Пользователи", "/admin/users"),
                (target.email, None),
            ],
            "target": target,
            "favorites": [{"fav": row[0], "car": row[1]} for row in favs],
        },
    )


@router.get("/admin/notifications", response_class=None)
def admin_notifications_page(
    request: Request,
    user: User | None = Depends(get_current_user),
    db: Session = Depends(get_db),
    user_id: int | None = None,
):
    if user is None:
        return RedirectResponse(url="/login?next=" + quote("/admin/notifications"), status_code=302)
    if not user.is_admin:
        return RedirectResponse(url="/", status_code=302)

    stats = AdminService(db).overview_stats()
    sent_stmt = (
        select(Notification, User)
        .join(User, User.id == Notification.user_id)
        .order_by(Notification.created_at.desc())
        .limit(100)
    )
    sent_rows = db.execute(sent_stmt).all()
    sent = [{"notif": row[0], "user": row[1]} for row in sent_rows]

    target_user = None
    target_favorites: list = []
    if user_id:
        target_user = db.get(User, user_id)
        if target_user:
            target_favorites = list(
                db.execute(
                    select(Car)
                    .join(Favorite, Favorite.car_id == Car.id)
                    .where(Favorite.user_id == target_user.id)
                    .order_by(Favorite.created_at.desc())
                    .limit(20)
                )
                .scalars()
                .all()
            )

    templates = request.app.state.templates
    return templates.TemplateResponse(
        "admin/notifications.html",
        {
            "request": request,
            "user": user,
            "page_title": "Сообщения",
            "page_subtitle": (
                "Отправляйте предложения зарегистрированным пользователям. "
                "Сообщение появляется в их личном кабинете и сохраняется как история."
            ),
            "breadcrumbs": [
                ("Админка", "/admin"),
                ("Сообщения", None),
            ],
            "stats": stats,
            "sent": sent,
            "target_user": target_user,
            "target_favorites": target_favorites,
        },
    )


@router.post("/admin/notifications/send")
def admin_notifications_send(
    request: Request,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
    user_id: int = Form(...),
    title: str = Form(""),
    body: str = Form(""),
    car_ids: str = Form(""),
):
    target = db.get(User, user_id)
    if target is None:
        return _admin_redirect(error="Пользователь не найден", path="/admin/notifications")
    body_clean = (body or "").strip()
    if not body_clean:
        return _admin_redirect(
            error="Текст сообщения не может быть пустым",
            path=f"/admin/notifications?user_id={user_id}",
        )

    attached: list[int] = []
    for piece in (car_ids or "").replace(",", " ").split():
        try:
            attached.append(int(piece))
        except ValueError:
            continue
    if attached:
        # Drop ids that no longer point at a real car so the inbox view
        # never renders broken thumbnails.
        existing = {
            cid
            for (cid,) in db.execute(
                select(Car.id).where(Car.id.in_(attached))
            ).all()
        }
        attached = [cid for cid in attached if cid in existing]

    try:
        NotificationService(db).send(
            user_id=target.id,
            sender=user,
            title=title,
            body=body_clean,
            attached_car_ids=attached,
        )
    except ValueError as exc:
        return _admin_redirect(
            error=str(exc),
            path=f"/admin/notifications?user_id={user_id}",
        )
    return _admin_redirect(
        f"Сообщение отправлено: {target.email}",
        path=f"/admin/users/{user_id}",
    )


@router.get("/admin/users.xlsx")
def admin_users_export_xlsx(
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
    q: str = "",
    verified: str = "",
    period: str = "",
):
    """Export the users table (with current filter state) as an XLSX."""

    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    fav_count_subq = (
        select(Favorite.user_id, func.count(Favorite.id).label("fav_count"))
        .group_by(Favorite.user_id)
        .subquery()
    )
    stmt = (
        select(User, func.coalesce(fav_count_subq.c.fav_count, 0).label("fav_count"))
        .outerjoin(fav_count_subq, fav_count_subq.c.user_id == User.id)
        .order_by(User.created_at.desc())
    )
    if q:
        like = f"%{q.strip()}%"
        stmt = stmt.where(
            User.email.ilike(like) | User.full_name.ilike(like) | User.phone.ilike(like)
        )
    if verified == "yes":
        stmt = stmt.where(
            User.email_verified_at.is_not(None) | User.phone_verified_at.is_not(None)
        )
    elif verified == "no":
        stmt = stmt.where(
            User.email_verified_at.is_(None) & User.phone_verified_at.is_(None)
        )
    if period:
        from datetime import datetime as _dt, timedelta as _td

        days = {"7": 7, "30": 30, "90": 90}.get(period)
        if days:
            stmt = stmt.where(User.created_at >= _dt.utcnow() - _td(days=days))

    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    headers = ["ID", "Email", "Имя", "Телефон", "Зарегистрирован", "Email подтверждён", "Телефон подтверждён", "Админ", "Избранных машин"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1f2533")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    for row in db.execute(stmt).all():
        u = row._mapping["User"]
        fav_count = int(row._mapping["fav_count"] or 0)
        ws.append([
            u.id,
            u.email,
            u.full_name or "",
            u.phone or "",
            u.created_at.strftime("%Y-%m-%d %H:%M") if u.created_at else "",
            "да" if u.email_verified_at else "нет",
            "да" if u.phone_verified_at else "нет",
            "да" if u.is_admin else "нет",
            fav_count,
        ])

    widths = [6, 30, 26, 18, 18, 18, 18, 8, 14]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"users-{datetime.utcnow().strftime('%Y%m%d-%H%M')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/admin/api/cars/search", response_class=None)
def admin_cars_search(
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
    q: str = "",
    limit: int = 20,
):
    """Lightweight typeahead for the admin compose form.

    Matches by exact id (when ``q`` is numeric) or by brand/model/variant
    substring. Returns just enough to render a label + thumbnail in the UI.
    """

    from ..utils.thumbs import resolve_thumbnail_url

    q_clean = (q or "").strip()
    stmt = select(Car).limit(max(1, min(int(limit or 20), 50)))
    if q_clean.isdigit():
        stmt = stmt.where(Car.id == int(q_clean))
    elif q_clean:
        like = f"%{q_clean}%"
        stmt = stmt.where(
            Car.brand.ilike(like)
            | Car.model.ilike(like)
            | Car.variant.ilike(like)
        ).order_by(Car.id.desc())
    else:
        stmt = stmt.order_by(Car.id.desc())

    cars = list(db.execute(stmt).scalars().all())
    out = []
    for car in cars:
        thumb = resolve_thumbnail_url(
            getattr(car, "thumbnail_url", None),
            getattr(car, "thumbnail_local_path", None),
        )
        label_parts = [car.brand or "", car.model or ""]
        title = " ".join(p for p in label_parts if p).strip() or f"#{car.id}"
        out.append({
            "id": car.id,
            "title": title,
            "subtitle": car.variant or "",
            "year": car.year,
            "thumbnail_url": thumb or "/static/img/no-photo.svg",
        })
    return {"results": out}


@router.get("/admin/analytics", response_class=None)
def admin_analytics_page(
    request: Request,
    user: User | None = Depends(get_current_user),
    db: Session = Depends(get_db),
    days: int = 30,
):
    if user is None:
        return RedirectResponse(url="/login?next=" + quote("/admin/analytics"), status_code=302)
    if not user.is_admin:
        return RedirectResponse(url="/", status_code=302)

    admin_svc = AdminService(db)
    stats = admin_svc.overview_stats()
    traffic = admin_svc.traffic_overview(days=days)
    recent = admin_svc.recent_users(days=traffic["days"], limit=20)
    cars_total = CarsService(db).total_cars()

    templates = request.app.state.templates
    return templates.TemplateResponse(
        "admin/analytics.html",
        {
            "request": request,
            "user": user,
            "page_title": "Аналитика",
            "page_subtitle": (
                "Посещения сайта (после согласия на cookies), регистрации "
                "и активность пользователей."
            ),
            "breadcrumbs": [
                ("Админка", "/admin"),
                ("Аналитика", None),
            ],
            "stats": stats,
            "cars_total": cars_total,
            "traffic": traffic,
            "recent_users": recent,
        },
    )


@router.get("/admin/contacts")
def admin_contacts_redirect(user: User | None = Depends(get_current_user)):
    if user is None:
        return RedirectResponse(url="/login?next=" + quote("/admin#tab=contacts"), status_code=302)
    return RedirectResponse(url="/admin#tab=contacts", status_code=302)


@router.get("/admin/recommended")
def admin_recommended_redirect(user: User | None = Depends(get_current_user)):
    if user is None:
        return RedirectResponse(url="/login?next=" + quote("/admin#tab=recommended"), status_code=302)
    return RedirectResponse(url="/admin#tab=recommended", status_code=302)


@router.get("/admin/calculator")
def admin_calculator_redirect(user: User | None = Depends(get_current_user)):
    if user is None:
        return RedirectResponse(url="/login?next=" + quote("/admin#tab=calculator"), status_code=302)
    return RedirectResponse(url="/admin#tab=calculator", status_code=302)


@router.get("/admin/customs")
def admin_customs_redirect(user: User | None = Depends(get_current_user)):
    if user is None:
        return RedirectResponse(url="/login?next=" + quote("/admin#tab=customs"), status_code=302)
    return RedirectResponse(url="/admin#tab=customs", status_code=302)


@router.get("/admin/home")
def admin_home_redirect(user: User | None = Depends(get_current_user)):
    if user is None:
        return RedirectResponse(url="/login?next=" + quote("/admin#tab=home"), status_code=302)
    return RedirectResponse(url="/admin#tab=home", status_code=302)


@router.post("/admin/contacts")
def update_contacts(
    request: Request,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
    lead_email: str = Form(""),
    contact_phone: str = Form(""),
    contact_email: str = Form(""),
    contact_address: str = Form(""),
    contact_tg: str = Form(""),
    contact_max: str = Form(""),
    contact_vk: str = Form(""),
    contact_avito: str = Form(""),
    contact_autoru: str = Form(""),
    contact_wa: str = Form(""),
    contact_ig: str = Form(""),
    contact_map_link: str = Form(""),
):
    admin_svc = AdminService(db)
    admin_svc.set_site_content(
        {
            "lead_email": lead_email,
            "contact_phone": contact_phone,
            "contact_email": contact_email,
            "contact_address": contact_address,
            "contact_tg": contact_tg,
            "contact_max": contact_max,
            "contact_vk": contact_vk,
            "contact_avito": contact_avito,
            "contact_autoru": contact_autoru,
            "contact_wa": contact_wa,
            "contact_ig": contact_ig,
            "contact_map_link": contact_map_link,
        }
    )
    return _admin_redirect("Контакты сохранены")


@router.post("/admin/recommended")
def update_recommended(
    request: Request,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
    max_age_years: int = Form(DEFAULT_CONFIG["max_age_years"]),
    price_min: int = Form(DEFAULT_CONFIG["price_min"]),
    price_max: int = Form(DEFAULT_CONFIG["price_max"]),
    mileage_max: int = Form(DEFAULT_CONFIG["mileage_max"]),
    reg_year_min: int = Form(DEFAULT_CONFIG["reg_year_min"]),
    reg_year_max: int = Form(DEFAULT_CONFIG["reg_year_max"]),
    power_hp_max: int = Form(DEFAULT_CONFIG["power_hp_max"]),
    engine_cc_max: int = Form(DEFAULT_CONFIG["engine_cc_max"]),
):
    save_config(
        {
            "max_age_years": max_age_years,
            "price_min": price_min,
            "price_max": price_max,
            "mileage_max": mileage_max,
            "reg_year_min": reg_year_min,
            "reg_year_max": reg_year_max,
            "power_hp_max": power_hp_max,
            "engine_cc_max": engine_cc_max,
        }
    )
    # Drop the cached SSR home payload so the new config takes effect on
    # the next refresh — otherwise the recommended block can stay stale.
    _bump_home_caches()
    return _admin_redirect("Параметры рекомендуемых сохранены")


@router.post("/admin/featured")
def update_featured(
    request: Request,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
    placement: str = Form("recommended"),
    car_ids: str = Form(""),
):
    from ..utils.featured_template import parse_featured_template

    ids = parse_featured_template(car_ids)
    AdminService(db).set_featured(placement, ids)
    _bump_home_caches()
    msg = f"Закреплено машин: {len(ids)}" if ids else "Список очищен — используется автоподбор"
    return _admin_redirect(msg)


@router.get("/admin/featured/template")
def download_featured_template(
    request: Request,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    from ..utils.featured_template import build_featured_template

    items = AdminService(db).list_featured("recommended")
    content = build_featured_template(items)
    return Response(
        content,
        media_type="text/plain; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=featured_template.txt"},
    )


@router.post("/admin/featured/upload")
async def upload_featured_template(
    request: Request,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
    file: UploadFile = File(...),
):
    from ..utils.featured_template import parse_featured_template

    if not file.filename.lower().endswith((".txt", ".csv")):
        return _admin_redirect(error="Неподдерживаемый формат — нужен .txt или .csv")
    raw = (await file.read()).decode("utf-8", errors="ignore")
    ids = parse_featured_template(raw)
    AdminService(db).set_featured("recommended", ids)
    _bump_home_caches()
    return _admin_redirect(f"Загружено {len(ids)} ID")


@router.get("/admin/customs/template")
def download_customs_template(
    request: Request,
    user: User = Depends(require_admin),
):
    data = load_customs_dict()
    content = build_util_template(data)
    return Response(
        content,
        media_type="text/plain",
        headers={"Content-Disposition": "attachment; filename=util_template.txt"},
    )


@router.post("/admin/customs/upload")
async def upload_customs_template(
    request: Request,
    user: User = Depends(require_admin),
    file: UploadFile = File(...),
):
    if not file.filename.lower().endswith((".txt", ".csv")):
        return _admin_redirect(error="Неподдерживаемый формат — нужен .txt или .csv")
    raw = (await file.read()).decode("utf-8", errors="ignore")
    data = load_customs_dict()
    stats = apply_util_template(data, raw)
    bump_customs_version(data)
    save_customs_dict(data)
    reset_customs_config_cache()
    msg = f"Обновлено строк: {stats.updated}"
    if stats.errors:
        msg += f", ошибок: {stats.errors}"
    return _admin_redirect(msg)


@router.post("/admin/customs/edit")
def edit_customs_row(
    request: Request,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
    age_bucket: str = Form("under_3"),
    cc_table: str = Form(""),
    power_type: str = Form("kw"),
    range_from: float = Form(...),
    range_to: float = Form(...),
    price_rub: float = Form(...),
):
    line = f"{age_bucket},{cc_table},{power_type},{range_from},{range_to},{price_rub}"
    data = load_customs_dict()
    stats = apply_util_template(data, line)
    bump_customs_version(data)
    save_customs_dict(data)
    reset_customs_config_cache()
    if stats.errors:
        return _admin_redirect(error=f"Не удалось применить строку (ошибок: {stats.errors})")
    return _admin_redirect("Строка утильсбора обновлена")


@router.post("/admin/home_content")
async def update_home_content(
    request: Request,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    form = await request.form()

    def get_value(name: str) -> str:
        raw = form.get(name)
        return raw.strip() if isinstance(raw, str) else ""

    def lines(name: str) -> list[str]:
        value = get_value(name)
        return [line.strip() for line in value.splitlines() if line.strip()]

    home = default_home_content()
    home["search"]["title"] = get_value("search_title")
    home["search"]["subtitle_suffix"] = get_value("search_subtitle_suffix")
    home["search"]["submit_label"] = get_value("search_submit_label")
    home["search"]["submit_suffix"] = get_value("search_submit_suffix")
    home["search"]["reset_label"] = get_value("search_reset_label")
    home["search"]["catalog_label"] = get_value("search_catalog_label")
    home["hero"]["stats_suffix"] = get_value("hero_stats_suffix")
    home["hero"]["title"] = get_value("hero_title")
    home["hero"]["subtitle"] = get_value("hero_subtitle")
    home["hero"]["note"] = get_value("hero_note")
    home["hero"]["benefits"] = lines("hero_benefits")
    home["hero"]["actions"]["primary_label"] = get_value("hero_primary_label")
    home["hero"]["actions"]["secondary_label"] = get_value("hero_secondary_label")
    home["hero"]["why_title"] = get_value("hero_why_title")
    home["hero"]["why_items"] = lines("hero_why_items")
    home["cases"]["title"] = get_value("cases_title")
    home["cases"]["subtitle"] = get_value("cases_subtitle")
    home["about"]["eyebrow"] = get_value("about_eyebrow")
    home["about"]["title"] = get_value("about_title")
    home["about"]["subtitle"] = get_value("about_subtitle")
    home["about"]["cards"][0]["title"] = get_value("about_card1_title")
    home["about"]["cards"][0]["items"] = lines("about_card1_items")
    home["about"]["cards"][1]["title"] = get_value("about_card2_title")
    home["about"]["cards"][1]["items"] = lines("about_card2_items")
    home["about"]["cards"][2]["title"] = get_value("about_card3_title")
    home["about"]["cards"][2]["items"] = lines("about_card3_items")
    home["about"]["cards"][3]["title"] = get_value("about_card4_title")
    home["about"]["cards"][3]["items"] = lines("about_card4_items")
    home["about"]["cards"][3]["paragraph"] = get_value("about_card4_paragraph")
    home["about"]["cards"][3]["emphasis"] = get_value("about_card4_emphasis")
    home["recommended"]["title"] = get_value("recommended_title")
    home["recommended"]["subtitle"] = get_value("recommended_subtitle")
    home["recommended"]["badge_label"] = get_value("recommended_badge_label")
    home["recommended"]["empty_note"] = get_value("recommended_empty_note")
    home["vehicle_types"]["title"] = get_value("vehicle_types_title")
    home["vehicle_types"]["subtitle"] = get_value("vehicle_types_subtitle")
    home["vehicle_types"]["empty_note"] = get_value("vehicle_types_empty_note")
    home["advantages"]["title"] = get_value("advantages_title")
    home["advantages"]["subtitle"] = get_value("advantages_subtitle")
    home["advantages"]["cards"][0]["title"] = get_value("advantages_card1_title")
    home["advantages"]["cards"][0]["text"] = get_value("advantages_card1_text")
    home["advantages"]["cards"][1]["title"] = get_value("advantages_card2_title")
    home["advantages"]["cards"][1]["text"] = get_value("advantages_card2_text")
    home["advantages"]["cards"][2]["title"] = get_value("advantages_card3_title")
    home["advantages"]["cards"][2]["text"] = get_value("advantages_card3_text")
    home["brands"]["title"] = get_value("brands_title")
    home["brands"]["subtitle"] = get_value("brands_subtitle")
    home["brands"]["empty_note"] = get_value("brands_empty_note")
    home["how_it_works"]["title"] = get_value("how_title")
    home["how_it_works"]["subtitle"] = get_value("how_subtitle")
    home["how_it_works"]["steps"][0]["title"] = get_value("how_step1_title")
    home["how_it_works"]["steps"][0]["text"] = get_value("how_step1_text")
    home["how_it_works"]["steps"][1]["title"] = get_value("how_step2_title")
    home["how_it_works"]["steps"][1]["text"] = get_value("how_step2_text")
    home["how_it_works"]["steps"][2]["title"] = get_value("how_step3_title")
    home["how_it_works"]["steps"][2]["text"] = get_value("how_step3_text")
    home["how_it_works"]["steps"][3]["title"] = get_value("how_step4_title")
    home["how_it_works"]["steps"][3]["text"] = get_value("how_step4_text")
    home["seo"]["title"] = get_value("seo_title")
    home["seo"]["paragraphs"][0] = get_value("seo_p1")
    home["seo"]["paragraphs"][1] = get_value("seo_p2")
    home["seo"]["paragraphs"][2] = get_value("seo_p3")
    home["lead"]["title"] = get_value("lead_title")
    home["lead"]["subtitle"] = get_value("lead_subtitle")
    home["lead"]["submit_label"] = get_value("lead_submit_label")
    home["lead"]["note"] = get_value("lead_note")
    home["lead"]["privacy_prefix"] = get_value("lead_privacy_prefix")
    home["lead"]["privacy_link_text"] = get_value("lead_privacy_link_text")
    home["contacts"]["title"] = get_value("contacts_title")
    home["contacts"]["subtitle_suffix"] = get_value("contacts_subtitle_suffix")
    home["contacts"]["callback_label"] = get_value("contacts_callback_label")
    home["contacts"]["hours_note"] = get_value("contacts_hours_note")
    home["contacts"]["call_button"] = get_value("contacts_call_button")
    home["contacts"]["email_button"] = get_value("contacts_email_button")

    admin_svc = AdminService(db)
    admin_svc.set_site_content({"home_content": serialize_home_content(home)})
    return _admin_redirect("Тексты главной страницы сохранены")


@router.post("/admin/calculator/upload")
async def upload_calculator_config(
    request: Request,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
    file: UploadFile = File(...),
):
    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        return _admin_redirect(error="Нужен файл .xlsx / .xlsm / .xls")
    try:
        with tempfile.NamedTemporaryFile(delete=False) as tmp:
            tmp.write(await file.read())
            tmp_path = Path(tmp.name)
        payload = CalculatorExtractor(tmp_path).extract()
        svc = CalculatorConfigService(db)
        svc.create(payload=payload, source="upload_xlsx", comment=f"upload {file.filename}")
    except Exception:
        return _admin_redirect(error="Не удалось прочитать файл — проверьте структуру шаблона")
    return _admin_redirect(f"Файл «{file.filename}» загружен, новая версия калькулятора сохранена")
